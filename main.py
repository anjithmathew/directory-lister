""" program for listing all directory as well as converting into excel sheets"""

import os
from openpyxl import Workbook


class ListFiles:
    def __init__(self):
        self.directory_name = os.getcwd()
        # printing current directory where the script running
        print("Current Directory:", self.directory_name)

    # lopp throught all the directory using os module
    def loop_through_all_files(self):
        self.list_file = []
        for root, dirs, files in os.walk(self.directory_name):
            for file in files:
                file_path = os.path.join(root, file)
                file_name = os.path.basename(file)
                file_size = os.path.getsize(file)
                # appending a tuple to the end of each []
                self.list_file.append((file_name, file_path, file_size))

        return self.list_file

    def convert_to_excel(self):
        wb = Workbook()
        ws = wb.active
        # title of excel Sheet
        ws.title = "file list"
        # title of excel blocks
        ws['A1'] = "file Name"
        ws['B1'] = "file Path"
        ws['C1'] = "file Size"
        # sorting the data
        self.sorted_list = sorted(
            self.list_file, key=lambda x: x[2],reverse=True)
        
        for idx,data in enumerate(self.sorted_list,start=2):
            ws.cell(row=idx, column=1, value=data[0])  # Col1 value
            ws.cell(row=idx, column=2, value=data[1])  # Col2 value
            ws.cell(row=idx, column=2, value=data[2])  # Col3 value
        output_filename = 'file_list.xlsx'
        wb.save(output_filename)
            


            


file_list = ListFiles()
print(file_list.loop_through_all_files())
file_list.convert_to_excel()
